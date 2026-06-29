# -*- coding: utf-8 -*-
"""导出全部测试用例为 Excel。运行: python test/scripts/export-testcases-xlsx.py"""
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / 'test' / 'docs' / '中试服务管理系统-测试用例汇总.xlsx'

HEADERS = [
    '用例编号', '用例分类', '模块', '子模块/页面', '用例类型', '优先级',
    '角色/端', '前置条件', '测试步骤', '预期结果', '自动化', '来源文档',
    '执行日期', '执行人', '测试结果', '缺陷编号', '备注'
]

rows = []

def add(cat, module, page, typ, pri, role, pre, step, expect, auto='', src='系统测试文档', cid='', note=''):
    rows.append([cid, cat, module, page, typ, pri, role, pre, step, expect, auto, src, '', '', '', '', note])


# ── 权限 AUTH ──
for r in [
    ('AUTH-01','未登录访问业务页','—','清除 token 访问 /enterprise/home','重定向 /login','P0'),
    ('AUTH-02','调度员登录','dispatcher','dispatcher/123456','进入 /center/home','P0'),
    ('AUTH-03','审核员登录','auditor','auditor/123456','进入 /center/home','P0'),
    ('AUTH-04','企业登录','enterprise','enterprise/123456','进入 /enterprise/home','P0'),
    ('AUTH-05','技术人员登录','technician','technician/123456','进入 /technician/home','P0'),
    ('AUTH-06','错误密码','任意','错误密码登录','提示失败不跳转','P1'),
    ('AUTH-07','企业访问中心路由','enterprise','访问 /center/dispatch/demand/workbench','跳转 /common/forbidden','P0'),
    ('AUTH-08','技术访问企业路由','technician','访问 /enterprise/demand/submit','跳转 forbidden','P0'),
    ('AUTH-09','调度访问审核页','dispatcher','访问 /center/audit/demand/verify','跳转 forbidden','P1'),
    ('AUTH-10','Token 无效','任意','篡改 token 刷新','应跳转登录(BUG-09)','P2'),
    ('AUTH-M01','小程序企业登录','enterprise/小程序','login 页登录','跳转企业首页','P0'),
    ('AUTH-M02','小程序未登录','小程序','清除 Storage 打开业务页','重定向登录','P0'),
    ('AUTH-M03','小程序技术登录','technician/小程序','technician 登录','跳转技术首页','P0'),
]:
    add('权限测试','公共',r[1],'功能',r[4],r[2],'—',r[3],r[4],'', '分工/系统测试文档', r[0])

# ── 公共 COM ──
for r in [
    ('COM-01','管理工作台首页','dispatcher','登录查看首页','待办卡片有数据','P0'),
    ('COM-02','企业门户首页','enterprise','登录查看','待办与步骤概览','P0'),
    ('COM-03','技术人员首页','technician','登录查看','任务待办列表','P0'),
    ('COM-04','待办办理跳转','各角色','点击待办办理','携带 projectId 正确跳转','P0'),
    ('COM-05','消息中心','任意','/common/messages','通知列表可已读','P1'),
    ('COM-06','个人中心','任意','/common/profile','信息展示可改密码','P1'),
    ('COM-07','五模块步骤条','enterprise','打开进度详情','与 stage 一致','P1'),
    ('COM-08','Web附件上传','enterprise','需求填报上传','返回 fileUrl','P1'),
    ('COM-09','小程序附件上传','enterprise/小程序','需求填报','无上传入口 BUG-10','P2'),
    ('COM-10','项目编号生成','enterprise','新建需求','ZS-YYYY-NNN','P1'),
]:
    add('公共功能','公共',r[1],'功能',r[4],r[2],'—',r[3],r[4],'', '分工/系统测试文档', r[0])

# ── 需求模块 D (系统测试文档) ──
demand_d = [
    ('D-01','需求预览确认','enterprise','DRAFT','预览页确认','可进入填报','P0'),
    ('D-02','需求填报提交','enterprise','DRAFT','填标题内容材料提交','SUBMITTED 调度待办+1','P0'),
    ('D-03','需求暂存','enterprise','DRAFT','保存草稿','保持 DRAFT','P2'),
    ('D-04','受理登记','dispatcher','SUBMITTED','工作台受理登记','ACCEPTING','P0'),
    ('D-05','材料审核通过','auditor','ACCEPTING','材料齐全同意受理(一步)','ACCEPTED','P0'),
    ('D-06','材料不齐','auditor','ACCEPTING','材料不齐退回','RETURNED','P1'),
    ('D-07','调度退回','dispatcher','SUBMITTED','退回意见','RETURNED','P0'),
    ('D-08','材料补充','enterprise','RETURNED','补充重新提交','SUBMITTED','P0'),
    ('D-09','同意受理','auditor','ACCEPTING','同意受理','ACCEPTED 企业待办签收','P0'),
    ('D-10','不予受理','auditor','ACCEPTING','不同意受理','stage=CLOSED','P1'),
    ('D-11','回执签收','enterprise','ACCEPTED','签收回执','RECEIPTED','P0'),
    ('D-12','需求归档','dispatcher','RECEIPTED','需求归档','ARCHIVED stage=EVALUATION','P0'),
    ('D-13','受理进度','enterprise','需求阶段','进度详情页','时间线步骤条正确','P1'),
    ('D-14','必填校验','enterprise','DRAFT','空标题内容提交','前端提示','P1'),
    ('D-15','越权提交','dispatcher','—','调度员 submit API','403','P1'),
]
for r in demand_d:
    add('模块功能','中试需求管理',r[1],'功能',r[5],r[2],r[3],r[4],r[5],'', '分工/系统测试文档', r[0])

# ── 需求详细 DM-UI/API/EX/NF (需求管理测试用例.md) ──
dm_ui = [
    ('DM-UI-001','需求预览','新建草稿','仅填标题保存','DRAFT+projectNo','P0'),
    ('DM-UI-002','需求预览','电话校验','非11位电话','前端提示','P1'),
    ('DM-UI-003','需求预览','材料上传','上传1份','列表显示','P0'),
    ('DM-UI-004','需求预览','进入填报','去填报','跳转 submit','P0'),
    ('DM-UI-005','需求预览','步骤条','草稿打开','节点=需求信息确认','P2'),
    ('DM-UI-010','需求填报','无projectId','直接访问 submit','重定向 preview','P1'),
    ('DM-UI-011','需求填报','空内容','内容空提交','提示不提交','P0'),
    ('DM-UI-012','需求填报','正常提交','填内容材料提交','SUBMITTED','P0'),
    ('DM-UI-013','需求填报','后端电话','无电话提交API','业务错误','P1'),
    ('DM-UI-020','受理工作台','待办分页','>10条翻页','数据正确','P2'),
    ('DM-UI-021','受理工作台','办理详情','点击办理','概要材料完整','P0'),
    ('DM-UI-022','受理工作台','受理登记','SUBMITTED登记','ACCEPTING','P0'),
    ('DM-UI-023','受理工作台','跳转退回','退回企业补充','跳转 reject','P1'),
    ('DM-UI-024','受理工作台','状态限制','ACCEPTING详情','无登记按钮','P1'),
    ('DM-UI-025','受理工作台','归档','RECEIPTED归档','ARCHIVED','P0'),
    ('DM-UI-026','受理工作台','越权','enterprise访问','403','P0'),
    ('DM-UI-030','退回意见','空意见','不填退回','提示必填','P1'),
    ('DM-UI-031','退回意见','正常退回','填原因确认','RETURNED','P0'),
    ('DM-UI-032','退回意见','非法状态','ACCEPTING reject','业务错误','P2'),
    ('DM-UI-040','材料审核','待办','审核员进入','仅 ACCEPTING','P0'),
    ('DM-UI-041','材料审核','预览','办理项目','附件可预览','P1'),
    ('DM-UI-042','材料审核','不齐无意见','材料不齐','提示填意见','P1'),
    ('DM-UI-043','材料审核','不齐提交','填意见不齐','RETURNED','P0'),
    ('DM-UI-044','材料审核','同意','材料齐全同意','ACCEPTED','P0'),
    ('DM-UI-045','材料审核','不同意','不同意受理','CLOSED','P1'),
    ('DM-UI-046','材料审核','缺参数','complete无accepted','业务错误','P2'),
    ('DM-UI-050','进度详情','流程日志','打开项目','时间线正确','P1'),
    ('DM-UI-051','进度详情','签收区','ACCEPTED','显示签收按钮','P0'),
    ('DM-UI-052','进度详情','签收','确认签收','RECEIPTED','P0'),
    ('DM-UI-053','进度详情','补充区','RETURNED','补充上传','P0'),
    ('DM-UI-054','进度详情','无材料补充','无材料提交','提示至少一份','P1'),
    ('DM-UI-055','进度详情','五模块进度','归档后','评估进行中','P2'),
    ('DM-UI-060','我的项目','空列表','无项目','提示发起','P2'),
    ('DM-UI-061','我的项目','阶段筛选','评估中','仅EVALUATION','P1'),
    ('DM-UI-062','我的项目','关键词','搜编号','匹配正确','P1'),
]
for r in dm_ui:
    add('模块功能(详细)','中试需求管理',r[1],'功能',r[5],'enterprise/dispatcher/auditor','—',r[3],r[4],'', 'test/docs/需求管理测试用例.md', r[0])

dm_api = [
    ('DM-API-01','POST','/api/projects/demand/create','enterprise','正常创建','200 DRAFT'),
    ('DM-API-02','POST','/api/projects/demand/create','dispatcher','越权','403'),
    ('DM-API-03','POST','.../demand/submit','enterprise','缺content','业务错误'),
    ('DM-API-04','POST','.../demand/submit','enterprise','错电话','业务错误'),
    ('DM-API-05','POST','.../accept-register','dispatcher','SUBMITTED','ACCEPTING'),
    ('DM-API-06','POST','.../verify','auditor','complete=false','RETURNED'),
    ('DM-API-07','POST','.../verify','auditor','accepted=true','ACCEPTED'),
    ('DM-API-08','POST','.../verify','auditor','accepted=false','CLOSED'),
    ('DM-API-09','POST','.../reject','dispatcher','SUBMITTED','RETURNED'),
    ('DM-API-10','POST','.../supplement','enterprise','RETURNED','SUBMITTED'),
    ('DM-API-11','POST','.../receipt','enterprise','ACCEPTED','RECEIPTED'),
    ('DM-API-12','POST','.../archive','dispatcher','RECEIPTED','ARCHIVED'),
    ('DM-API-13','GET','/api/demand/todos','各角色','分页','按角色过滤'),
    ('DM-API-14','GET','/api/demand/enterprise/projects','enterprise','筛选搜索','分页正确'),
    ('DM-API-15','POST','/api/files/upload','enterprise','multipart','fileUrl'),
    ('DM-API-16','GET','/api/files/download','未登录','下载','401/403'),
]
for r in dm_api:
    add('接口测试','中试需求管理',r[2],'API','P1',r[3],'—',r[4],r[5],'', 'test/docs/需求管理测试用例.md', r[0])

for r in [
    ('DM-EX-01','重复提交','非DRAFT submit','业务错误'),
    ('DM-EX-02','跳过登记','SUBMITTED verify','业务错误'),
    ('DM-EX-03','跳过审核','ACCEPTING receipt','业务错误'),
    ('DM-EX-04','未签收归档','ACCEPTED archive','业务错误'),
    ('DM-EX-05','跨企业','A访问B项目','403'),
    ('DM-EX-06','非法ID','不存在projectId','404/错误'),
    ('DM-EX-07','并发提交','两次submit','第二次失败'),
]:
    add('异常边界','中试需求管理',r[1],'异常','P2','—','—',r[2],r[3],'', 'test/docs/需求管理测试用例.md', r[0])

for r in [
    ('DM-NF-01','性能','工作台加载','<2s'),
    ('DM-NF-02','易用','待办办理','带projectId'),
    ('DM-NF-03','易用','上传Toast','明确反馈'),
    ('DM-NF-04','兼容','附件预览','多格式正常'),
    ('DM-NF-05','安全','下载鉴权','须JWT'),
]:
    add('非功能','中试需求管理',r[2],'非功能','P2','—','—',r[2],r[3],'', 'test/docs/需求管理测试用例.md', r[0])

# ── 评估 E ──
for r in [
    ('E-01','评估前置核查','dispatcher','EVALUATION','核查条件具备','进入条件评估','P0'),
    ('E-02','条件不具备','dispatcher','前置','判定不具备','企业待办补充','P0'),
    ('E-03','条件材料补充','enterprise','退回','补充提交','回前置核查','P0'),
    ('E-04','整改通知','dispatcher','条件否','发通知','企业收到','P2'),
    ('E-05','资源核定','dispatcher','条件具备','核定','进可行性','P1'),
    ('E-06','可行性通过','auditor','—','审查可行','待办结论','P0'),
    ('E-07','可行性否','auditor','—','不可行','企业补充','P1'),
    ('E-08','评估材料补充','enterprise','退回','补充','回可行性','P1'),
    ('E-09','评估结论','auditor','可行','出具通过','企业签收','P0'),
    ('E-10','结论签收','enterprise','—','签收','进反馈','P0'),
    ('E-11','评估反馈','enterprise','已签收','提交意见','可归档','P1'),
    ('E-12','评估归档','dispatcher','—','归档','stage=DISPATCH','P0'),
    ('E-13','结论详情','enterprise','—','只读查看','无编辑','P1'),
]:
    add('模块功能','中试评估管理',r[1],'功能',r[5],r[2],r[3],r[4],r[5],'', '分工/系统测试文档', r[0])

# ── 调度 S ──
for r in [
    ('S-01','资源匹配','dispatcher','DISPATCH','匹配成功','可派发','P0'),
    ('S-02','匹配失败','dispatcher','MATCH','失败','MATCH_FAILED','P1'),
    ('S-03','任务派发','dispatcher','匹配成功','选资源技术','技术待办','P0'),
    ('S-04','派单通知','dispatcher','ASSIGNED','通知','PENDING_RECEIVE','P0'),
    ('S-05','派单后待办','dispatcher','PENDING_RECEIVE','查待办','应有待办 BUG-01','P0'),
    ('S-06','任务接收','technician','PENDING_RECEIVE','接收','RECEIVED','P0'),
    ('S-07','确认签收','technician','RECEIVED','确认','CONFIRMED','P0'),
    ('S-08','进度填报','technician','CONFIRMED','填报100%','EXECUTING','P0'),
    ('S-09','进度查看','enterprise','EXECUTING','只读查看','无填报','P0'),
    ('S-10','进度督办','dispatcher','EXECUTING','督办','可查看','P1'),
    ('S-11','异常重派','dispatcher','异常','重派','技术再接收','P1'),
    ('S-12','执行确认','dispatcher','100%','确认完成','EXEC_DONE','P0'),
    ('S-13','调度归档','dispatcher','EXEC_DONE','归档','stage=FEEDBACK','P0'),
    ('S-14','跳过进度','dispatcher','CONFIRMED','直接确认','应拒绝 BUG-06','P2'),
]:
    add('模块功能','中试调度管理',r[1],'功能',r[5],r[2],r[3],r[4],r[5],'', '分工/系统测试文档', r[0])

# ── 反馈 F ──
for r in [
    ('F-01','结果提交','technician','FEEDBACK','提交','待审核','P0'),
    ('F-02','数据校验','technician','已提交','校验','进审核','P1'),
    ('F-03','报告审核通过','auditor','待审','通过','待复核','P0'),
    ('F-04','报告退回','auditor','待审','退回','待修改','P0'),
    ('F-05','结果修改','technician','退回','修改提交','回审核','P0'),
    ('F-06','复核确认','auditor','通过','复核通过','企业反馈','P0'),
    ('F-07','复核拒绝','auditor','—','复核不通过','BUG-02强制通过','P2'),
    ('F-08','报告归档','auditor','复核通过','归档','进反馈审核','P1'),
    ('F-09','复核通知','auditor','—','通知','企业可见','P2'),
    ('F-10','复核意见反馈','enterprise','—','提交','待反馈审核','P0'),
    ('F-11','反馈审核','auditor','已反馈','通过','可案例归档','P0'),
    ('F-12','反馈审核退回','auditor','—','退回','企业重反馈','P1'),
    ('F-13','案例归档','auditor','审核通过','归档','stage=ARCHIVE','P0'),
    ('F-14','复核详情','enterprise','—','只读','显示结论','P1'),
]:
    add('模块功能','中试反馈管理',r[1],'功能',r[5],r[2],r[3],r[4],r[5],'', '分工/系统测试文档', r[0])

# ── 档案 A ──
for r in [
    ('A-01','台账维护','auditor','ARCHIVE','标记完整','LEDGER_OK','P0'),
    ('A-02','台账不完整','auditor','PENDING','不完整','停留维护','P0'),
    ('A-03','档案确认','auditor','LEDGER_OK','确认','进归集','P1'),
    ('A-04','资料归集','auditor','CONFIRMED','归集','COLLECTED','P1'),
    ('A-05','归集后待办','auditor','COLLECTED','查待办','无错误待办 BUG-03','P1'),
    ('A-06','周期统计','dispatcher','COLLECTED','打开统计','图表有数据','P0'),
    ('A-07','成功率分析','dispatcher','—','打开分析','渲染正常','P1'),
    ('A-08','统计不可用','dispatcher','不足','标记不可用','回台账','P2'),
    ('A-09','简报生成','dispatcher','STATS_OK','生成','BRIEF_GENERATED','P0'),
    ('A-10','简报审核','auditor','BRIEF_GENERATED','通过','企业可看','P0'),
    ('A-11','简报查看','enterprise','BRIEF_PUBLISHED','查看','只读','P0'),
    ('A-12','档案归档','auditor','BRIEF_PUBLISHED','归档','CLOSED','P0'),
]:
    add('模块功能','中试档案管理',r[1],'功能',r[5],r[2],r[3],r[4],r[5],'', '分工/系统测试文档', r[0])

# ── 非功能 NF ──
for r in [
    ('NF-01','性能','首页待办','<2s','P1'),
    ('NF-02','性能','统计图表','3s出图','P1'),
    ('NF-03','兼容','Chrome/Edge','布局正常','P1'),
    ('NF-04','兼容','微信开发者工具','小程序正常','P1'),
    ('NF-05','易用','待办办理','一键跳转','P1'),
    ('NF-06','易用','登录演示账号','快捷填充','P2'),
    ('NF-07','可靠','H2持久化','重启不丢','P1'),
    ('NF-08','可靠','重复操作','拒绝或幂等','P2'),
]:
    add('非功能','全系统',r[2],'非功能',r[4],'—','—',r[2],r[3],'', '分工/系统测试文档', r[0])

# ── 业务流程场景（逐步） ──
flow_cases = {
    'DM-FLOW-001': ('中试需求管理','主路径','P0','新建或清空H2','npm run demand:flow(至RECEIPTED)', [
        (1,'enterprise','预览填表上传保存','DRAFT'),
        (2,'enterprise','填报提交','SUBMITTED'),
        (3,'dispatcher','工作台核对受理登记','ACCEPTING'),
        (4,'auditor','材料审核同意','ACCEPTED'),
        (5,'enterprise','进度页签收','RECEIPTED'),
        (6,'dispatcher','工作台归档','ARCHIVED/EVALUATION'),
        (7,'enterprise','我的项目查进度','日志完整'),
    ]),
    'DM-FLOW-002': ('中试需求管理','否支↺','P0','已提交','', [
        (1,'enterprise','提交','SUBMITTED'),
        (2,'dispatcher','退回','RETURNED'),
        (3,'enterprise','补充重提','SUBMITTED'),
        (4,'—','继续主路径','闭环'),
    ]),
    'DM-FLOW-003': ('中试需求管理','否支↺','P1','ACCEPTING','', [
        (1,'auditor','材料不齐','RETURNED'),
        (2,'enterprise','补充','SUBMITTED'),
        (3,'—','继续主路径','闭环'),
    ]),
    'DM-FLOW-004': ('中试需求管理','终止','P1','ACCEPTING','', [
        (1,'auditor','不同意受理','CLOSED'),
        (2,'enterprise','已结案列表','可见不可操作'),
    ]),
    'DM-FLOW-005': ('中试需求管理','查询','P1','多项目','', [
        (1,'enterprise','打开我的项目','分页列表'),
        (2,'enterprise','阶段筛选','过滤正确'),
        (3,'enterprise','关键词搜索','匹配'),
    ]),
    'DM-FLOW-006': ('中试需求管理','双端联调','P1','—','', [
        (1,'小程序','企业提交','Toast成功'),
        (2,'Web','调度待办','同步'),
        (4,'小程序','签收','状态一致'),
    ]),
    'TC-D1': ('中试需求管理','主路径','P0','企业登录','手工E2E', [
        (1,'enterprise','预览→填报→提交','SUBMITTED'),
        (2,'dispatcher','受理登记','ACCEPTING'),
        (3,'auditor','材料审核同意(一步)','ACCEPTED'),
        (4,'enterprise','回执签收','RECEIPTED'),
        (5,'dispatcher','需求归档','EVALUATION'),
        (6,'enterprise','我的项目查进度','日志可查'),
    ]),
    'TC-D2': ('中试需求管理','否支↺','P0','已提交','', [
        (1,'enterprise','提交','SUBMITTED'),
        (2,'dispatcher','退回','RETURNED'),
        (3,'enterprise','补充','SUBMITTED'),
        (4,'—','继续TC-D1','闭环'),
    ]),
    'TC-E1': ('中试评估管理','主路径','P0','stage=EVALUATION','', [
        (1,'dispatcher','前置核查','—'),
        (2,'dispatcher','可行性可行','—'),
        (3,'auditor','评估结论','—'),
        (4,'enterprise','签收反馈归档','DISPATCH'),
    ]),
    'TC-E2': ('中试评估管理','否支↺','P1','—','', [
        (1,'dispatcher','条件不具备','企业补充'),
        (2,'enterprise','补充','回前置'),
        (3,'—','继续TC-E1','闭环'),
    ]),
    'TC-S1': ('中试调度管理','主路径','P0','DISPATCH','', [
        (1,'dispatcher','匹配派发','技术待办'),
        (2,'technician','接收签收进度100%','—'),
        (3,'dispatcher','执行确认','FEEDBACK'),
        (4,'enterprise','进度只读','—'),
    ]),
    'TC-S2': ('中试调度管理','否支↺','P1','—','', [
        (1,'dispatcher','异常重派','—'),
        (2,'dispatcher','重派','技术再接收'),
        (3,'technician','完成S1','恢复'),
    ]),
    'TC-F1': ('中试反馈管理','主路径','P0','FEEDBACK','', [
        (1,'technician','提交','待审'),
        (2,'auditor','审核复核','企业反馈'),
        (3,'enterprise','反馈意见','ARCHIVE'),
    ]),
    'TC-F2': ('中试反馈管理','否支↺','P0','—','', [
        (1,'technician','提交','—'),
        (2,'auditor','退回','修改'),
        (3,'technician','修改','回审核'),
    ]),
    'TC-A1': ('中试档案管理','主路径','P0','ARCHIVE','', [
        (1,'auditor','台账确认归集','—'),
        (2,'dispatcher','统计简报','—'),
        (3,'auditor','简报审核归档','CLOSED'),
    ]),
    'TC-A2': ('中试档案管理','否支↺','P1','—','', [
        (1,'auditor','台账不完整','停留'),
        (2,'auditor','补全','可统计'),
    ]),
    'TC-X2': ('全系统','双端联调','P1','—','', [
        (1,'小程序','企业提交','—'),
        (2,'Web','调度待办','同步'),
        (4,'小程序','技术进度','Web同步'),
    ]),
}
for cid, (mod, typ, pri, pre, auto, steps) in flow_cases.items():
    for sn, role, step, exp in steps:
        add('业务流程', mod, cid, '场景', pri, role, pre, f'步骤{sn}: {step}', exp, auto, '业务流程测试用例/需求管理测试用例', f'{cid}-S{sn}')

# TC-X1 验收检查点
for r in [
    ('TC-X1-C1','最终stage','CLOSED'),
    ('TC-X1-C2','项目编号','ZS-2026-NNN'),
    ('TC-X1-C3','四角色待办','无阻塞待办'),
    ('TC-X1-C4','步骤条','五模块完成'),
    ('TC-X1-C5','workflow_log','关键节点有记录'),
]:
    add('业务流程','全系统','TC-X1 P0主链','验收检查','P0','—','单项目全流程','检查'+r[0],r[1], '手工E2E', '业务流程测试用例', r[0])

# ── 缺陷 BUG ──
bugs = [
    ('BUG-01','P0','调度','dispatcherTodos缺PENDING_RECEIVE','已复现'),
    ('BUG-16','P0','调度','dispatcherTodos缺RECEIVED','已复现'),
    ('BUG-02','P1','反馈','review硬编码passed=true','已复现'),
    ('BUG-03','P1','档案','auditorTodos含COLLECTED','已复现'),
    ('BUG-04','P1','调度','enterpriseTodos缺PROGRESS_ACKED','已复现'),
    ('BUG-05','P1','通用','步骤条ordinal错位','代码确认'),
    ('BUG-06','P2','调度','execConfirm可跳过进度','代码确认'),
    ('BUG-07','P2','反馈','passed null语义不一','代码确认'),
    ('BUG-08','P2','安全','uploads公开','已复现'),
    ('BUG-09','P2','前端','code401未处理','代码确认'),
    ('BUG-10','P2','小程序','无文件上传','代码确认'),
    ('BUG-11','P2','需求','项目编号非年度','代码确认'),
    ('BUG-17','P1','调度','progressPct NPE','已复现'),
    ('BUG-18','P2','测试','单元测试8/12','已复现'),
    ('BUG-19','P2','Web','build TS错误','已复现'),
    ('BUG-12','P3','评估','conclusion无log','代码确认'),
    ('BUG-13','P3','调度','NOTICED未使用','代码确认'),
    ('BUG-14','P3','安全','无PreAuthorize','已知'),
    ('BUG-15','P3','安全','IDOR可读任意项目','已知'),
]
for b in bugs:
    add('缺陷跟踪', b[2], b[0], '缺陷', b[1], '—', '—', '代码审查/实测', b[3], '', '分工/系统测试文档', b[0], b[4])

# ── 单元测试 ──
for r in [
    ('UT-D-01','DemandServiceTest','submit_movesDraftToSubmitted','提交'),
    ('UT-D-02','DemandServiceTest','supplement_movesReturnedBackToSubmitted','退回补充'),
    ('UT-E-01','EvaluationServiceTest','conditionFail_movesToReturned','条件否'),
    ('UT-E-02','EvaluationServiceTest','conditionSupplement_loopsBackToPrecheck','补充回前置'),
    ('UT-S-01','DispatchServiceTest','execConfirm_keepsDispatchStage','执行确认'),
    ('UT-S-02','DispatchServiceTest','archive_movesToFeedbackStage','调度归档'),
    ('UT-F-01','FeedbackServiceTest','auditReject_returnsToModify','审核否'),
    ('UT-F-02','FeedbackServiceTest','caseArchive_movesToArchiveStage','案例归档'),
    ('UT-A-01','ArchiveServiceTest','updateLedger_complete_movesToConfirm','台账'),
    ('UT-A-02','ArchiveServiceTest','archive_movesToClosedStage','结案'),
]:
    add('单元测试', r[1].replace('ServiceTest',''), r[0], '单元', 'P1', '—', '—', f'mvn test {r[2]}', '断言通过', 'mvn test', 'server/src/test', r[0])


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = '测试用例汇总'

    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    widths = [14, 12, 14, 18, 8, 6, 14, 20, 36, 28, 14, 22, 10, 8, 8, 10, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{len(rows)+1}'

    # 统计页
    ws2 = wb.create_sheet('统计')
    ws2.append(['用例分类', '数量'])
    from collections import Counter
    cnt = Counter(r[1] for r in rows)
    for k, v in sorted(cnt.items(), key=lambda x: -x[1]):
        ws2.append([k, v])
    ws2.append(['合计', len(rows)])
    ws2.append([])
    ws2.append(['说明', '汇总来源：分工/系统测试文档.md、test/docs/需求管理测试用例.md、原型图/业务流程测试用例.md'])
    ws2.append(['生成日期', '2026-06-26'])

    wb.save(OUT)
    return len(rows)


if __name__ == '__main__':
    n = build_workbook()
    print(f'已生成 {n} 条测试用例 -> {OUT}')
